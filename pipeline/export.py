"""统一导出：把「话术转录 + 评论区弹幕 + 直播数据」按房间汇总成可读文件。

数据源（两条线解耦，各写各的库，导出时按 room/live id 关联）：
  transcripts.db   音频转写（话术口播）   —— recorder_rotate + transcribe_batch
  multi_events.db  弹幕/评论/进场/直播数据 —— 当前监听后端写入

输出（exports/，整目录 .gitignore 忽略）：
  <id>.md      单房间：话术全文 + 评论区样本 + 直播数据曲线
  summary.csv  所有房间一行汇总，便于后续打分

用法：
  python -m pipeline.export                 # 导出全部房间
  python -m pipeline.export <房间号>          # 只导出指定房间
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import sqlite3
import sys
import time
from bisect import bisect_left
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from . import anchor_profiles, config, license_manager


ROOMS_JSON = config.ROOMS_JSON


def _conn(path) -> sqlite3.Connection | None:
    """打开只读连接；库不存在返回 None（某条线可能还没跑）。"""
    if not path.exists():
        return None
    c = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    c.row_factory = sqlite3.Row
    return c


@dataclass(frozen=True)
class TranscriptRow:
    room_id: str
    segment_ts: int
    duration_sec: float | None
    text: str
    char_count: int
    mp3_name: str          # 相对录音名 <房间号>/<seq>.mp3，便于按导出文字回放录音
    segment_id: int | None = None
    recording_status: str | None = None
    capture_start: float | None = None
    capture_end: float | None = None
    speaker_label: str = ""
    speaker_similarity: float | None = None
    speaker_change: str = ""


@dataclass(frozen=True)
class RecordingTimelineRow:
    id: int
    room_id: str
    seq: int | None
    kind: str
    status: str
    file_path: str | None
    capture_start: float | None
    capture_end: float | None
    duration_sec: float | None
    file_size: int | None
    error: str | None
    transcribed: bool
    transcript_preview: str


@dataclass(frozen=True)
class RoomBundle:
    """单房间导出所需的全部数据。音频按 mp3 文件名里的 room_id，弹幕按 live_id。"""

    rid: str
    nickname: str          # 主播昵称（room_meta），区分各房间用
    transcripts: list[TranscriptRow]
    timeline: list[RecordingTimelineRow]
    chats: list[tuple[str, str]]          # (user_name, content)
    stats: list[tuple[int, int, int]]      # (ts, current_online, total_pv)
    event_counts: dict[str, int]
    source_rid: str = ""
    session_id: str = ""
    session_day: str = ""


@dataclass(frozen=True)
class SpeakerLabel:
    room_id: str
    file_name: str
    label: str
    similarity: float | None
    change_status: str


def _fmt_ts(ts: int) -> str:
    # 弹幕事件用毫秒时间戳（13位），转写段用秒（10位）；统一成秒再格式化。
    if ts > 1_000_000_000_000:
        ts //= 1000
    try:
        return datetime.fromtimestamp(ts).strftime("%H:%M:%S")
    except (ValueError, OSError, OverflowError):
        return str(ts)


def _fmt_dt(ts: int) -> str:
    """完整日期时间（24/7 跨天，需带日期才能区分）。毫秒/秒戳自动归一。"""
    if ts > 1_000_000_000_000:
        ts //= 1000
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, OSError, OverflowError):
        return str(ts)


def _speaker_cn(label: str) -> str:
    """speaker_A -> 发言人A；低置信度/无标签使用清晰中文状态。"""
    label = (label or "").strip()
    if label.startswith("speaker_") and label != "speaker_uncertain":
        suffix = label.removeprefix("speaker_")
        return f"发言人{suffix}"
    if label == "speaker_uncertain":
        return "待确认"
    return ""


def _speaker_change_cn(status: str) -> str:
    """仅向用户展示已经确认的换人点，其余内部状态不进入导出表。"""
    if status in {"change_confirmed_start", "change_confirmed"}:
        return "换人"
    return ""


def _room_from_audio_path(value: str) -> str:
    parts = Path(value.replace("\\", "/")).parts
    try:
        index = parts.index("audio")
        return parts[index + 1]
    except (ValueError, IndexError):
        return ""


def load_speaker_labels() -> dict[tuple[str, str], SpeakerLabel]:
    """Read historical CSV labels, then let automatic labels override them."""
    path = config.SPEAKER_LABELS_CSV
    labels: dict[tuple[str, str], SpeakerLabel] = {}
    if path.exists():
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                for row in csv.DictReader(handle):
                    file_name = Path(row.get("file_name", "")).name
                    room_id = _room_from_audio_path(row.get("file_path", ""))
                    if not room_id or not file_name:
                        continue
                    raw_similarity = row.get("similarity_prev", "")
                    try:
                        similarity = float(raw_similarity) if raw_similarity else None
                    except ValueError:
                        similarity = None
                    labels[(room_id, file_name)] = SpeakerLabel(
                        room_id=room_id,
                        file_name=file_name,
                        label=_speaker_cn(row.get("speaker_label", "")),
                        similarity=similarity,
                        change_status=row.get("change_status", ""),
                    )
        except (OSError, csv.Error):
            pass
    c = _conn(config.SPEAKER_DB_PATH)
    if c is not None:
        try:
            for row in c.execute(
                "SELECT room_id,file_name,speaker_label,similarity,change_status "
                "FROM speaker_labels"
            ):
                labels[(row["room_id"], row["file_name"])] = SpeakerLabel(
                    room_id=row["room_id"],
                    file_name=row["file_name"],
                    label=_speaker_cn(row["speaker_label"]),
                    similarity=row["similarity"],
                    change_status=row["change_status"] or "",
                )
        except sqlite3.Error:
            pass
        finally:
            c.close()
    return labels


def _parse_stat(content: str) -> tuple[int, int]:
    """'current=123;total_pv=4567' -> (123, 4567)；解析失败回退 (0,0)。"""
    cur = pv = 0
    for part in content.split(";"):
        k, _, v = part.partition("=")
        if k == "current" and v.isdigit():
            cur = int(v)
        elif k == "total_pv" and v.isdigit():
            pv = int(v)
    return cur, pv


def load_transcripts(
    rid: str, speaker_labels: dict[tuple[str, str], SpeakerLabel] | None = None
) -> list[TranscriptRow]:
    c = _conn(config.DB_PATH)
    if c is None:
        return []
    try:
        try:
            rows = c.execute(
                "SELECT room_id, segment_ts, duration_sec, text, char_count, mp3_name, "
                "segment_id, recording_status, capture_start, capture_end "
                "FROM transcripts WHERE room_id = ? ORDER BY segment_ts",
                (rid,),
            ).fetchall()
        except sqlite3.Error:
            try:
                rows = c.execute(
                    "SELECT room_id, segment_ts, duration_sec, text, char_count, mp3_name "
                    "FROM transcripts WHERE room_id = ? ORDER BY segment_ts",
                    (rid,),
                ).fetchall()
            except sqlite3.Error:
                return []
        labels = speaker_labels if speaker_labels is not None else load_speaker_labels()
        out = []
        for r in rows:
            file_name = Path((r["mp3_name"] or "").replace("\\", "/")).name
            speaker = labels.get((rid, file_name))
            out.append(TranscriptRow(
                r["room_id"], r["segment_ts"], r["duration_sec"],
                r["text"], r["char_count"], r["mp3_name"],
                r["segment_id"] if "segment_id" in r.keys() else None,
                r["recording_status"] if "recording_status" in r.keys() else None,
                r["capture_start"] if "capture_start" in r.keys() else None,
                r["capture_end"] if "capture_end" in r.keys() else None,
                speaker.label if speaker else "",
                speaker.similarity if speaker else None,
                speaker.change_status if speaker else "",
            ))
        return out
    finally:
        c.close()


def load_recording_timeline(rid: str) -> list[RecordingTimelineRow]:
    c = _conn(config.DB_PATH)
    if c is None:
        return []
    try:
        try:
            rows = c.execute(
                "SELECT rt.*, tr.text AS transcript_text, tr.id AS transcript_id "
                "FROM recording_timeline rt "
                "LEFT JOIN transcripts tr ON tr.segment_id = rt.id "
                "WHERE rt.room_id = ? "
                "ORDER BY COALESCE(rt.capture_start, rt.created_ts), rt.id",
                (rid,),
            ).fetchall()
        except sqlite3.Error:
            return []
        out: list[RecordingTimelineRow] = []
        for r in rows:
            text = r["transcript_text"] or ""
            preview = (text[:60] + "…") if len(text) > 60 else text
            out.append(RecordingTimelineRow(
                r["id"],
                r["room_id"],
                r["seq"],
                r["kind"],
                r["status"],
                r["file_path"],
                r["capture_start"],
                r["capture_end"],
                r["duration_sec"],
                r["file_size"],
                r["error"],
                r["transcript_id"] is not None,
                preview,
            ))
        return out
    finally:
        c.close()


def load_room_meta() -> dict[str, str]:
    """读 room_meta 的 {live_id: nickname}。老库无此表时返回空。"""
    c = _conn(config.EVENTS_DB)
    if c is None:
        return {}
    try:
        try:
            rows = c.execute("SELECT live_id, nickname FROM room_meta").fetchall()
        except sqlite3.Error:
            return {}
        return {r["live_id"]: (r["nickname"] or "") for r in rows}
    finally:
        c.close()


def monitored_room_ids() -> list[str]:
    """当前控制台监听清单，兼容旧字符串列表和新版主播资料对象列表。"""
    if not ROOMS_JSON.exists():
        return []
    try:
        rows = json.loads(ROOMS_JSON.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return []
    if not isinstance(rows, list):
        return []
    ids: set[str] = set()
    for row in rows:
        rid = row.get("rid") if isinstance(row, dict) else row
        value = str(rid or "").strip()
        if value and value.isdigit():
            ids.add(value)
    return sorted(ids)


def configured_room_meta() -> dict[str, str]:
    """Read display names saved by the console, without touching network state."""
    return {
        rid: str(profile.get("anchor_name") or "").strip()
        for rid, profile in configured_room_profiles().items()
        if str(profile.get("anchor_name") or "").strip()
    }


def configured_room_profiles() -> dict[str, dict[str, str]]:
    """Read saved anchor name/avatar/profile metadata without touching network state."""
    out = anchor_profiles.load_profiles()
    if not ROOMS_JSON.exists():
        return out
    try:
        rows = json.loads(ROOMS_JSON.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return out
    if not isinstance(rows, list):
        return out
    for row in rows:
        if not isinstance(row, dict):
            continue
        rid = str(row.get("rid") or "").strip()
        if not rid.isdigit():
            continue
        cached = out.get(rid, {})
        out[rid] = {
            "anchor_name": str(row.get("anchor_name") or "").strip() or cached.get("anchor_name"),
            "avatar_url": cached.get("avatar_url") or str(row.get("avatar_url") or "").strip(),
            "source_url": str(row.get("source_url") or "").strip() or cached.get("source_url"),
            "sec_user_id": str(row.get("sec_user_id") or "").strip() or cached.get("sec_user_id"),
        }
    return out


def room_display_names() -> dict[str, str]:
    """Merge runtime-discovered names with console metadata."""
    names = load_room_meta()
    names.update(configured_room_meta())
    return names


def safe_export_stem(nickname: str, rid: str) -> str:
    """Create a Windows-safe nickname filename, falling back to the live id."""
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", (nickname or "").strip())
    name = name.rstrip(" .")[:80]
    return name or str(rid)


def room_export_filename(rid: str, suffix: str = ".xlsx") -> str:
    """Filename for a single-room export; nickname first, live id fallback."""
    return safe_export_stem(room_display_names().get(str(rid), ""), str(rid)) + suffix


def unique_bundle_stems(bundles: list["RoomBundle"]) -> dict[str, str]:
    """Use nicknames where possible; append live id only when nicknames collide."""
    counts: dict[str, int] = {}
    raw: dict[str, str] = {}
    for bundle in bundles:
        stem = safe_export_stem(bundle.nickname, bundle.rid)
        raw[bundle.rid] = stem
        counts[stem.casefold()] = counts.get(stem.casefold(), 0) + 1
    return {
        bundle.rid: (
            f"{raw[bundle.rid]}_{bundle.rid}"
            if counts[raw[bundle.rid].casefold()] > 1
            else raw[bundle.rid]
        )
        for bundle in bundles
    }


def _write_conn(path) -> sqlite3.Connection | None:
    if not path.exists():
        return None
    return sqlite3.connect(str(path))


def cleanup_unmonitored_data(keep_ids: list[str] | None = None) -> dict[str, int]:
    """清理不在当前监听清单里的旧库数据，避免历史脏房间混入导出。

    只清 SQLite 中的 transcripts/events/room_meta；录音文件和历史 exports 不删，避免误删可回听素材。
    若监听清单为空则不清理，防止配置异常时把库清空。
    """
    keep = set(keep_ids or monitored_room_ids())
    if not keep:
        return {"transcripts": 0, "events": 0, "room_meta": 0}

    deleted = {"transcripts": 0, "recording_timeline": 0, "events": 0, "room_meta": 0}
    placeholders = ",".join("?" for _ in keep)

    tc = _write_conn(config.DB_PATH)
    if tc is not None:
        try:
            cur = tc.execute(f"DELETE FROM transcripts WHERE room_id NOT IN ({placeholders})", tuple(keep))
            deleted["transcripts"] = cur.rowcount if cur.rowcount != -1 else 0
            try:
                cur = tc.execute(f"DELETE FROM recording_timeline WHERE room_id NOT IN ({placeholders})", tuple(keep))
                deleted["recording_timeline"] = cur.rowcount if cur.rowcount != -1 else 0
            except sqlite3.Error:
                pass
            tc.commit()
        finally:
            tc.close()

    ec = _write_conn(config.EVENTS_DB)
    if ec is not None:
        try:
            cur = ec.execute(f"DELETE FROM events WHERE live_id NOT IN ({placeholders})", tuple(keep))
            deleted["events"] = cur.rowcount if cur.rowcount != -1 else 0
            try:
                cur = ec.execute(f"DELETE FROM room_meta WHERE live_id NOT IN ({placeholders})", tuple(keep))
                deleted["room_meta"] = cur.rowcount if cur.rowcount != -1 else 0
            except sqlite3.Error:
                pass
            ec.commit()
        finally:
            ec.close()

    return deleted


def cleanup_export_files(keep_ids: list[str]) -> int:
    """清理 exports/ 里不属于当前监听清单的单房间文件；汇总文件保留/覆盖。"""
    keep = set(keep_ids)
    if not config.EXPORT_DIR.exists():
        return 0
    removed = 0
    for p in config.EXPORT_DIR.iterdir():
        if not p.is_file() or p.suffix.lower() not in {".xlsx", ".md"}:
            continue
        if p.stem in {"汇总", "summary"}:
            continue
        if p.stem not in keep:
            try:
                p.unlink()
                removed += 1
            except OSError:
                pass
    return removed


def _event_range_clause(start_ts: float | None, end_ts: float | None) -> tuple[str, list[int]]:
    clauses: list[str] = []
    params: list[int] = []
    if start_ts is not None:
        clauses.append("ts >= ?")
        params.append(int(float(start_ts) * 1000))
    if end_ts is not None:
        clauses.append("ts < ?")
        params.append(int(float(end_ts) * 1000))
    if not clauses:
        return "", []
    return " AND " + " AND ".join(clauses), params


def load_events(
    rid: str,
    start_ts: float | None = None,
    end_ts: float | None = None,
) -> tuple[list[tuple[str, str]], list[tuple[int, int, int]], dict[str, int]]:
    """返回 (chat样本, stat序列, 各类事件计数)。弹幕库按 live_id 关联。"""
    c = _conn(config.EVENTS_DB)
    if c is None:
        return [], [], {}
    try:
        range_clause, range_params = _event_range_clause(start_ts, end_ts)
        counts = {
            r["event_type"]: r["n"]
            for r in c.execute(
                "SELECT event_type, COUNT(*) AS n FROM events "
                f"WHERE live_id = ?{range_clause} GROUP BY event_type",
                (rid, *range_params),
            )
        }
        chats = [
            (r["user_name"] or "", r["content"] or "")
            for r in c.execute(
                "SELECT user_name, content FROM events "
                f"WHERE live_id = ? AND event_type = 'chat'{range_clause} ORDER BY ts",
                (rid, *range_params),
            )
        ]
        stats = []
        for r in c.execute(
            "SELECT content, ts FROM events "
            f"WHERE live_id = ? AND event_type = 'stat'{range_clause} ORDER BY ts",
            (rid, *range_params),
        ):
            cur, pv = _parse_stat(r["content"] or "")
            stats.append((r["ts"], cur, pv))
        return chats, stats, counts
    finally:
        c.close()


def load_chats_ts(rid: str) -> list[tuple[int, str, str]]:
    """全部聊天弹幕，带毫秒时间戳，按时间排序。xlsx 弹幕表用。"""
    c = _conn(config.EVENTS_DB)
    if c is None:
        return []
    try:
        return [
            (r["ts"], r["user_name"] or "", r["content"] or "")
            for r in c.execute(
                "SELECT ts, user_name, content FROM events "
                "WHERE live_id = ? AND event_type = 'chat' ORDER BY ts",
                (rid,),
            )
        ]
    finally:
        c.close()


def all_room_ids() -> list[str]:
    """两条线里出现过的全部房间号并集。"""
    ids: set[str] = set()
    tc = _conn(config.DB_PATH)
    if tc is not None:
        try:
            ids.update(r[0] for r in tc.execute("SELECT DISTINCT room_id FROM transcripts"))
            try:
                ids.update(r[0] for r in tc.execute("SELECT DISTINCT room_id FROM recording_timeline"))
            except sqlite3.Error:
                pass
        finally:
            tc.close()
    ec = _conn(config.EVENTS_DB)
    if ec is not None:
        try:
            ids.update(r[0] for r in ec.execute("SELECT DISTINCT live_id FROM events"))
        finally:
            ec.close()
    if config.AUDIO_DIR.exists():
        ids.update(
            path.name
            for path in config.room_audio_dirs()
            if path.name.isdigit()
        )
    return sorted(ids)


def _transcript_counts_by_room() -> dict[str, int]:
    c = _conn(config.DB_PATH)
    if c is None:
        return {}
    try:
        try:
            return {
                str(r["room_id"]): int(r["n"] or 0)
                for r in c.execute(
                    "SELECT room_id, COUNT(*) AS n FROM transcripts GROUP BY room_id"
                )
            }
        except sqlite3.Error:
            return {}
    finally:
        c.close()


def _event_summary_counts_by_room() -> dict[str, dict[str, int]]:
    c = _conn(config.EVENTS_DB)
    if c is None:
        return {}
    out: dict[str, dict[str, int]] = {}
    try:
        try:
            rows = c.execute(
                "SELECT live_id, event_type, COUNT(*) AS n FROM events GROUP BY live_id, event_type"
            )
        except sqlite3.Error:
            return {}
        for r in rows:
            rid = str(r["live_id"])
            event_type = str(r["event_type"] or "")
            count = int(r["n"] or 0)
            item = out.setdefault(rid, {"events": 0, "chats": 0, "stats": 0})
            item["events"] += count
            if event_type == "chat":
                item["chats"] = count
            elif event_type == "stat":
                item["stats"] = count
        return out
    finally:
        c.close()


def _audio_summary_for_room(rid: str) -> tuple[int, int]:
    room_dir = config.AUDIO_DIR / rid
    if not room_dir.exists():
        return 0, 0
    count = 0
    total_bytes = 0
    for path in room_dir.glob("*.mp3"):
        if not path.is_file():
            continue
        try:
            total_bytes += path.stat().st_size
            count += 1
        except OSError:
            continue
    return count, total_bytes


def _date_label(ts: float | int | None) -> str:
    if ts is None:
        return ""
    try:
        return datetime.fromtimestamp(float(ts)).strftime("%Y-%m-%d")
    except (OSError, OverflowError, ValueError, TypeError):
        return ""


def _recording_sessions_for_room(rid: str) -> tuple[float, list[dict[str, object]]]:
    """Return lightweight day/session summaries for picker UI."""
    by_date: dict[str, dict[str, object]] = {}
    for row in load_recording_timeline(rid):
        if row.kind == "gap":
            continue
        start = row.capture_start
        end = row.capture_end
        duration = row.duration_sec
        if duration is None and start is not None and end is not None:
            duration = max(0.0, float(end) - float(start))
        duration = max(0.0, float(duration or 0))
        if duration <= 0:
            continue
        day = _date_label(start or end)
        if not day:
            continue
        item = by_date.setdefault(day, {"date": day, "duration_sec": 0.0, "segments": 0})
        item["duration_sec"] = float(item["duration_sec"]) + duration
        item["segments"] = int(item["segments"]) + 1
    if not by_date:
        for row in load_transcripts(rid):
            duration = max(0.0, float(row.duration_sec or 0))
            if duration <= 0:
                continue
            day = _date_label(row.capture_start or row.segment_ts)
            if not day:
                continue
            item = by_date.setdefault(day, {"date": day, "duration_sec": 0.0, "segments": 0})
            item["duration_sec"] = float(item["duration_sec"]) + duration
            item["segments"] = int(item["segments"]) + 1
    sessions = sorted(by_date.values(), key=lambda x: str(x["date"]), reverse=True)
    for idx, item in enumerate(sessions, start=1):
        item["session_id"] = f"{rid}:{item['date']}"
        item["label"] = f"{item['date']} 录制 {_format_duration(float(item['duration_sec']))}"
        item["index"] = idx
    total = sum(float(item["duration_sec"]) for item in sessions)
    return total, sessions


def _format_duration(seconds: float) -> str:
    seconds = max(0, int(round(seconds)))
    hours, rem = divmod(seconds, 3600)
    minutes, secs = divmod(rem, 60)
    if hours:
        return f"{hours}小时{minutes:02d}分"
    if minutes:
        return f"{minutes}分{secs:02d}秒"
    return f"{secs}秒"


def data_room_summaries() -> list[dict]:
    """列出所有历史数据房间及数据量，供控制台选择导出/清理。"""
    meta = room_display_names()
    profiles = anchor_profiles.load_profiles()
    transcript_counts = _transcript_counts_by_room()
    event_counts = _event_summary_counts_by_room()
    summaries = []
    for rid in all_room_ids():
        audio_files, audio_bytes = _audio_summary_for_room(rid)
        events = event_counts.get(rid, {})
        profile = profiles.get(rid, {})
        record_seconds, sessions = _recording_sessions_for_room(rid)
        nickname = meta.get(rid, "") or profile.get("anchor_name", "")
        summaries.append({
            "rid": rid,
            "nickname": nickname,
            "anchor_name": nickname,
            "avatar_url": profile.get("avatar_url", ""),
            "transcripts": transcript_counts.get(rid, 0),
            "events": events.get("events", 0),
            "chats": events.get("chats", 0),
            "stats": events.get("stats", 0),
            "audio_files": audio_files,
            "audio_bytes": audio_bytes,
            "record_seconds": record_seconds,
            "record_duration_text": _format_duration(record_seconds),
            "dates": [str(s["date"]) for s in sessions],
            "sessions": sessions,
        })
    return summaries


def delete_room_data(rid: str) -> dict[str, int]:
    """彻底删除一个房间的业务数据。仅由用户明确触发，导出流程绝不调用。"""
    rid = str(rid).strip()
    if not rid or not rid.isdigit():
        raise ValueError("房间号格式无效")
    deleted = {
        "transcripts": 0,
        "recording_timeline": 0,
        "events": 0,
        "room_meta": 0,
        "audio_files": 0,
        "export_files": 0,
        "speaker_labels": 0,
    }
    tc = _write_conn(config.DB_PATH)
    if tc is not None:
        try:
            for table in ("transcripts", "recording_timeline"):
                try:
                    cur = tc.execute(f"DELETE FROM {table} WHERE room_id = ?", (rid,))
                    deleted[table] = max(0, cur.rowcount)
                except sqlite3.Error:
                    pass
            tc.commit()
        finally:
            tc.close()
    ec = _write_conn(config.EVENTS_DB)
    if ec is not None:
        try:
            for table in ("events", "room_meta"):
                try:
                    key = "live_id"
                    cur = ec.execute(f"DELETE FROM {table} WHERE {key} = ?", (rid,))
                    deleted[table] = max(0, cur.rowcount)
                except sqlite3.Error:
                    pass
            ec.commit()
        finally:
            ec.close()
    sc = _write_conn(config.SPEAKER_DB_PATH)
    if sc is not None:
        try:
            try:
                cur = sc.execute("DELETE FROM speaker_labels WHERE room_id = ?", (rid,))
                deleted["speaker_labels"] = max(0, cur.rowcount)
                sc.execute("DELETE FROM speaker_profiles WHERE room_id = ?", (rid,))
            except sqlite3.Error:
                pass
            sc.commit()
        finally:
            sc.close()

    room_dir = (config.AUDIO_DIR / rid).resolve()
    audio_root = config.AUDIO_DIR.resolve()
    if room_dir.parent == audio_root and room_dir.exists():
        deleted["audio_files"] = sum(1 for p in room_dir.rglob("*") if p.is_file())
        shutil.rmtree(room_dir)
    if config.FAILED_DIR.exists():
        for path in config.FAILED_DIR.glob(f"{rid}__*"):
            if path.is_file():
                path.unlink()
                deleted["audio_files"] += 1

    if config.EXPORT_DIR.exists():
        for path in config.EXPORT_DIR.iterdir():
            if path.is_file() and path.stem == rid:
                path.unlink()
                deleted["export_files"] += 1
    return deleted


def build_bundle(
    rid: str,
    nickname: str = "",
    speaker_labels: dict[tuple[str, str], SpeakerLabel] | None = None,
    *,
    start_ts: float | None = None,
    end_ts: float | None = None,
    session_id: str = "",
    session_day: str = "",
) -> RoomBundle:
    transcripts = load_transcripts(rid, speaker_labels)
    timeline = load_recording_timeline(rid)
    if start_ts is not None or end_ts is not None:
        transcripts = [
            row for row in transcripts
            if _row_in_time_range(row.capture_start or row.segment_ts, start_ts, end_ts)
        ]
        timeline = [
            row for row in timeline
            if _row_in_time_range(row.capture_start or row.capture_end, start_ts, end_ts)
        ]
    chats, stats, counts = load_events(rid, start_ts, end_ts)
    return RoomBundle(
        rid,
        nickname,
        transcripts,
        timeline,
        chats,
        stats,
        counts,
        source_rid=rid,
        session_id=session_id,
        session_day=session_day,
    )


def _row_in_time_range(value: float | int | None, start_ts: float | None, end_ts: float | None) -> bool:
    if value is None:
        return False
    ts = float(value)
    if ts > 1_000_000_000_000:
        ts /= 1000
    if start_ts is not None and ts < float(start_ts):
        return False
    if end_ts is not None and ts >= float(end_ts):
        return False
    return True


def render_markdown(b: RoomBundle) -> str:
    title = f"# 房间 {b.rid}" + (f"（{b.nickname}）" if b.nickname else "")
    lines: list[str] = [title, ""]

    # 直播数据概览
    peak = max((s[1] for s in b.stats), default=0)
    last_pv = b.stats[-1][2] if b.stats else 0
    lines += [
        "## 直播数据",
        f"- 弹幕样本点: {len(b.stats)}",
        f"- 在线人数峰值: {peak}",
        f"- 累计看播次数（平台展示PV，非去重人数）: {last_pv}",
        "- 口径说明: 该值表示平台实时展示的直播间打开/看播次数，不等于进入直播间的独立人数；实时值可能因平台校准而波动或下降。",
        f"- 事件计数: " + (", ".join(f"{k}={v}" for k, v in sorted(b.event_counts.items())) or "无"),
        "",
    ]
    if b.stats:
        lines.append("在线人数曲线（采样）:")
        step = max(1, len(b.stats) // 20)
        for ts, cur, pv in b.stats[::step]:
            lines.append(f"  {_fmt_ts(ts)}  在线 {cur}  平台展示PV {pv}")
        lines.append("")

    # 话术转录
    total_chars = sum(t.char_count for t in b.transcripts)
    lines += ["## 话术转录", f"- 段数: {len(b.transcripts)}  总字数: {total_chars}", ""]
    for t in b.transcripts:
        dur = f"{t.duration_sec:.0f}s" if t.duration_sec is not None else "?"
        speaker = f", {t.speaker_label}" if t.speaker_label else ""
        lines.append(f"### 段 {_fmt_ts(t.segment_ts)}  ({dur}, {t.char_count}字{speaker}, 录音 {t.mp3_name})")
        lines.append(t.text or "（空）")
        lines.append("")

    # 评论区弹幕
    lines += ["## 评论区弹幕", f"- 条数: {len(b.chats)}", ""]
    for name, content in b.chats:
        lines.append(f"- {name}: {content}")
    lines.append("")

    return "\n".join(lines)


def export_room(rid: str, nickname: str = "") -> RoomBundle:
    b = build_bundle(rid, nickname)
    out = config.EXPORT_DIR / f"{rid}.md"
    out.write_text(render_markdown(b), encoding="utf-8")
    return b


# sheet 列头（单房间表；汇总表会在最前面再加一列「房间号」）
TRANSCRIPT_HEADER = [
    "捕获开始", "捕获结束", "录音状态", "时长(秒)", "字数",
    "话术内容", "发言人", "换人标记", "录音文件",
]
TIMELINE_HEADER = [
    "seq", "类型", "状态", "捕获开始", "捕获结束", "时长(秒)",
    "文件大小", "录音文件", "是否转写", "转写预览", "错误摘要",
]
CHAT_HEADER = ["时间", "用户", "内容"]
STAT_HEADER = ["时间", "在线人数", "累计看播次数（平台展示PV，非去重人数）"]
TOTAL_HEADER = [
    "房间号", "主播", "发言人", "捕获开始", "捕获结束", "时长(秒)",
    "话术内容", "字数", "录音状态", "录音文件", "同期弹幕数",
    "当时在线人数", "当时累计看播次数（平台展示PV，非去重人数）", "房间弹幕总数", "房间进场总数",
    "房间点赞总数", "房间在线峰值", "房间累计看播次数（平台展示PV，非去重人数）", "换人标记",
]
SPEAKER_SUMMARY_HEADER = [
    "房间号", "主播", "发言人", "话术段数", "累计时长(秒)", "话术字数",
    "首次发言", "最后发言", "确认换人次数",
]


def _transcript_rows(transcripts: list[TranscriptRow]) -> list[list]:
    """话术行：起始时间用录制起戳；录音文件列指回 audio/<房间号>/<seq>.mp3 便于回听。"""
    rows = []
    for t in transcripts:
        start = t.capture_start if t.capture_start is not None else t.segment_ts
        end = t.capture_end
        rows.append([
            _fmt_dt(int(start)),
            _fmt_dt(int(end)) if end is not None else "",
            t.recording_status or "",
            round(t.duration_sec, 1) if t.duration_sec is not None else "",
            t.char_count,
            t.text or "",
            t.speaker_label,
            _speaker_change_cn(t.speaker_change),
            t.mp3_name,
        ])
    return rows


def _timeline_rows(rows: list[RecordingTimelineRow]) -> list[list]:
    out = []
    for r in rows:
        out.append([
            r.seq if r.seq is not None else "",
            r.kind,
            r.status,
            _fmt_dt(int(r.capture_start)) if r.capture_start is not None else "",
            _fmt_dt(int(r.capture_end)) if r.capture_end is not None else "",
            round(r.duration_sec, 2) if r.duration_sec is not None else "",
            r.file_size if r.file_size is not None else "",
            r.file_path or "",
            "是" if r.transcribed else "否",
            r.transcript_preview,
            r.error or "",
        ])
    return out


def _chat_rows(chats: list[tuple[int, str, str]]) -> list[list]:
    return [[_fmt_dt(ts), name, content] for ts, name, content in chats]


def _stat_rows(stats: list[tuple[int, int, int]]) -> list[list]:
    return [[_fmt_dt(ts), cur, pv] for ts, cur, pv in stats]


def _epoch_seconds(ts: int | float) -> float:
    return float(ts) / 1000.0 if ts > 1_000_000_000_000 else float(ts)


def _nearest_stat(
    stats: list[tuple[int, int, int]], target_ts: float
) -> tuple[int | str, int | str]:
    """取距离话术开始时间最近的直播数据采样点。"""
    if not stats:
        return "", ""
    stat_seconds = [_epoch_seconds(row[0]) for row in stats]
    index = bisect_left(stat_seconds, target_ts)
    candidates = [i for i in (index - 1, index) if 0 <= i < len(stats)]
    best = min(candidates, key=lambda i: abs(stat_seconds[i] - target_ts))
    return stats[best][1], stats[best][2]


def _total_rows(
    bundles: list[RoomBundle],
    chats_provider=None,
) -> list[list]:
    """每段话术一行，同时关联发言人、同期弹幕与直播数据。

    chats_provider：可选 rid->[(ts,name,content)] 的取数函数；样本导出传入内存数据，
    避免读真实弹幕库。默认 None 时按原路径从 EVENTS_DB 读。
    """
    rows: list[list] = []
    for b in bundles:
        timed_chats = chats_provider(b.rid) if chats_provider else load_chats_ts(b.rid)
        chat_seconds = [_epoch_seconds(item[0]) for item in timed_chats]
        peak = max((stat[1] for stat in b.stats), default=0)
        last_pv = b.stats[-1][2] if b.stats else 0
        for t in b.transcripts:
            start = float(t.capture_start if t.capture_start is not None else t.segment_ts)
            end = float(
                t.capture_end
                if t.capture_end is not None
                else start + (t.duration_sec or 0)
            )
            chat_count = bisect_left(chat_seconds, end) - bisect_left(chat_seconds, start)
            current_online, current_pv = _nearest_stat(b.stats, start)
            rows.append([
                b.rid,
                b.nickname,
                t.speaker_label,
                _fmt_dt(int(start)),
                _fmt_dt(int(end)) if end else "",
                round(t.duration_sec, 1) if t.duration_sec is not None else "",
                t.text or "",
                t.char_count,
                t.recording_status or "",
                t.mp3_name,
                chat_count,
                current_online,
                current_pv,
                len(b.chats),
                b.event_counts.get("member", 0),
                b.event_counts.get("like", 0),
                peak,
                last_pv,
                _speaker_change_cn(t.speaker_change),
            ])
    return rows


def _speaker_summary_rows(bundles: list[RoomBundle]) -> list[list]:
    groups: dict[tuple[str, str], list[TranscriptRow]] = {}
    nicknames = {b.rid: b.nickname for b in bundles}
    for b in bundles:
        for transcript in b.transcripts:
            label = transcript.speaker_label or "未标注"
            groups.setdefault((b.rid, label), []).append(transcript)
    rows = []
    for (rid, label), transcripts in sorted(groups.items()):
        starts = [
            float(t.capture_start if t.capture_start is not None else t.segment_ts)
            for t in transcripts
        ]
        ends = [
            float(
                t.capture_end
                if t.capture_end is not None
                else (t.capture_start if t.capture_start is not None else t.segment_ts)
                + (t.duration_sec or 0)
            )
            for t in transcripts
        ]
        rows.append([
            rid,
            nicknames.get(rid, ""),
            label,
            len(transcripts),
            round(sum(t.duration_sec or 0 for t in transcripts), 1),
            sum(t.char_count for t in transcripts),
            _fmt_dt(int(min(starts))) if starts else "",
            _fmt_dt(int(max(ends))) if ends else "",
            sum(t.speaker_change == "change_confirmed_start" for t in transcripts),
        ])
    return rows


def _style_sheet(ws, widths: list[int], wrap_col: int | None = None) -> None:
    """统一样式：表头加粗、冻结首行、设列宽，可选指定列自动换行。"""
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if wrap_col:
        for row in ws.iter_rows(min_row=2, min_col=wrap_col, max_col=wrap_col):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")


def _export_watermark_rows(scope: str, room_ids: list[str]) -> list[list[str]]:
    status = license_manager.current_status()
    payload = status.payload if isinstance(status.payload, dict) else {}
    license_id = str(payload.get("license_id") or payload.get("activation_id") or status.mode or "unknown")
    return [
        ["授权ID", license_id],
        ["授权状态", status.mode],
        ["设备指纹", license_manager.current_device_hash()[:12]],
        ["软件版本", config.LICENSE_APP_VERSION],
        ["导出范围", scope],
        ["房间数量", str(len(room_ids))],
        ["房间号", ", ".join(sorted(str(rid) for rid in room_ids))],
        ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["说明", "此文件由直播复盘侠导出，含授权溯源信息，请勿外传。"],
    ]


def _apply_export_watermark(wb, *, scope: str, room_ids: list[str]) -> None:
    """Add traceable export metadata when the signed policy requires watermarking."""
    if license_manager.current_policy().get("export_watermark") is False:
        return
    title = _safe_sheet_title("授权水印", set(wb.sheetnames))
    ws = wb.create_sheet(title, 0)
    ws.append(["项目", "内容"])
    for row in _export_watermark_rows(scope, room_ids):
        ws.append(row)
    _style_sheet(ws, [18, 90], wrap_col=2)


def _safe_sheet_title(base: str, used: set[str]) -> str:
    """Excel sheet 名限制 31 字符且不能含特殊符号；重名时追加序号。"""
    title = re.sub(r"[\[\]\:\*\?\/\\]", "_", base).strip() or "未命名"
    title = title[:31]
    if title not in used:
        used.add(title)
        return title
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = f"{title[:31 - len(suffix)]}{suffix}"
        if candidate not in used:
            used.add(candidate)
            return candidate
    raise RuntimeError("无法生成唯一工作表名")


def build_workbook(rid: str):
    """单房间工作簿：话术(时间段+口播) / 弹幕(逐条+时间) / 直播数据(在线曲线)。

    三张表共用同一时间轴，便于「某时段主播说了什么、评论区在刷什么、在线掉没掉」对照看。
    """
    from openpyxl import Workbook

    speaker_labels = load_speaker_labels()
    transcripts = load_transcripts(rid, speaker_labels)
    timeline = load_recording_timeline(rid)
    chats = load_chats_ts(rid)
    room_chats, stats, counts = load_events(rid)
    bundle = RoomBundle(rid, room_display_names().get(rid, ""), transcripts, timeline,
                        room_chats, stats, counts)

    wb = Workbook()

    ws = wb.active
    ws.title = "话术"
    ws.append(TRANSCRIPT_HEADER)
    for r in _transcript_rows(transcripts):
        ws.append(r)
    _style_sheet(ws, [20, 20, 12, 9, 7, 90, 14, 14, 28], wrap_col=6)

    ws_all = wb.create_sheet("总表")
    ws_all.append(TOTAL_HEADER)
    for row in _total_rows([bundle]):
        ws_all.append(row)
    _style_sheet(
        ws_all,
        [14, 16, 14, 20, 20, 10, 90, 8, 12, 28, 12, 14, 14, 14, 14, 14, 14, 14, 20],
        wrap_col=7,
    )

    ws_speaker = wb.create_sheet("发言人汇总")
    ws_speaker.append(SPEAKER_SUMMARY_HEADER)
    for row in _speaker_summary_rows([bundle]):
        ws_speaker.append(row)
    _style_sheet(ws_speaker, [14, 16, 14, 12, 16, 12, 20, 20, 14])

    ws_tl = wb.create_sheet("录音时间轴")
    ws_tl.append(TIMELINE_HEADER)
    for r in _timeline_rows(timeline):
        ws_tl.append(r)
    _style_sheet(ws_tl, [8, 10, 10, 20, 20, 9, 10, 34, 9, 60, 40], wrap_col=10)

    ws_chat = wb.create_sheet("弹幕")
    ws_chat.append(CHAT_HEADER)
    for r in _chat_rows(chats):
        ws_chat.append(r)
    _style_sheet(ws_chat, [20, 22, 60])

    ws_stat = wb.create_sheet("直播数据")
    ws_stat.append(STAT_HEADER)
    for r in _stat_rows(stats):
        ws_stat.append(r)
    _style_sheet(ws_stat, [20, 12, 12])

    _apply_export_watermark(wb, scope="单房间导出", room_ids=[rid])
    return wb


def build_summary_workbook(bundles: list[RoomBundle]):
    """全房间总表：汇总(每房间一行) + 每主播话术独立 sheet + 弹幕(全房间)。

    话术不混在一张表里：每个监听房间单独一张「话术-主播/房间号」表，便于逐个评估；
    弹幕仍合并到一张表，并保留房间号/主播列用于筛选。
    """
    from openpyxl import Workbook

    wb = Workbook()
    used_titles: set[str] = set()

    ws = wb.active
    ws.title = "汇总"
    used_titles.add(ws.title)
    ws.append(SUMMARY_HEADER)
    for b in bundles:
        ws.append(summary_row(b))
    _style_sheet(ws, [16, 16, 10, 10, 16, 12, 10, 10, 10, 12, 12])

    ws_all = wb.create_sheet("总表")
    used_titles.add(ws_all.title)
    ws_all.append(TOTAL_HEADER)
    for row in _total_rows(bundles):
        ws_all.append(row)
    _style_sheet(
        ws_all,
        [14, 16, 14, 20, 20, 10, 90, 8, 12, 28, 12, 14, 14, 14, 14, 14, 14, 14, 20],
        wrap_col=7,
    )

    ws_speaker = wb.create_sheet("发言人汇总")
    used_titles.add(ws_speaker.title)
    ws_speaker.append(SPEAKER_SUMMARY_HEADER)
    for row in _speaker_summary_rows(bundles):
        ws_speaker.append(row)
    _style_sheet(ws_speaker, [14, 16, 14, 12, 16, 12, 20, 20, 14])

    for b in bundles:
        label = b.nickname or b.rid
        ws_tr = wb.create_sheet(_safe_sheet_title(f"话术-{label}", used_titles))
        ws_tr.append(["房间号", "主播"] + TRANSCRIPT_HEADER)
        for r in _transcript_rows(b.transcripts):
            ws_tr.append([b.rid, b.nickname] + r)
        _style_sheet(
            ws_tr,
            [14, 16, 20, 20, 12, 9, 7, 90, 14, 14, 28],
            wrap_col=8,
        )

    ws_tl = wb.create_sheet("录音时间轴")
    ws_tl.append(["房间号", "主播"] + TIMELINE_HEADER)
    for b in bundles:
        for r in _timeline_rows(b.timeline):
            ws_tl.append([b.rid, b.nickname] + r)
    _style_sheet(ws_tl, [14, 16, 8, 10, 10, 20, 20, 9, 10, 34, 9, 60, 40], wrap_col=12)

    ws_ch = wb.create_sheet("弹幕")
    ws_ch.append(["房间号", "主播"] + CHAT_HEADER)
    for b in bundles:
        for ts, name, content in load_chats_ts(b.rid):
            ws_ch.append([b.rid, b.nickname, _fmt_dt(ts), name, content])
    _style_sheet(ws_ch, [14, 16, 20, 22, 60])

    ws_stat = wb.create_sheet("直播数据")
    ws_stat.append(["房间号", "主播"] + STAT_HEADER)
    for b in bundles:
        for ts, cur, pv in b.stats:
            ws_stat.append([b.rid, b.nickname, _fmt_dt(ts), cur, pv])
    _style_sheet(ws_stat, [14, 16, 20, 12, 12])

    _apply_export_watermark(wb, scope="汇总导出", room_ids=[b.rid for b in bundles])
    return wb


def xlsx_bytes(rid: str) -> bytes:
    """单房间工作簿的字节流，供 Web 直接下载。"""
    import io
    buf = io.BytesIO()
    build_workbook(rid).save(buf)
    return buf.getvalue()


def summary_xlsx_bytes(bundles: list[RoomBundle]) -> bytes:
    """全房间总表的字节流，供 Web 直接下载。"""
    import io
    buf = io.BytesIO()
    build_summary_workbook(bundles).save(buf)
    return buf.getvalue()


def selected_xlsx_bytes(rids: list[str]) -> bytes:
    """按用户选择的房间生成一个完整总表，不写磁盘、不清理任何数据。"""
    ids = sorted({str(rid).strip() for rid in rids if str(rid).strip()})
    meta = room_display_names()
    speaker_labels = load_speaker_labels()
    bundles = [build_bundle(rid, meta.get(rid, ""), speaker_labels) for rid in ids]
    return summary_xlsx_bytes(bundles)


def write_xlsx(rid: str) -> "object":
    out = config.EXPORT_DIR / room_export_filename(rid)
    build_workbook(rid).save(out)
    return out


SUMMARY_HEADER = [
    "room_id", "主播", "发言人数", "话术段数", "话术时长(秒)", "话术字数",
    "弹幕条数", "进场数", "点赞数", "在线峰值", "累计看播次数（平台展示PV，非去重人数）",
]


def summary_row(b: RoomBundle) -> list:
    peak = max((s[1] for s in b.stats), default=0)
    last_pv = b.stats[-1][2] if b.stats else 0
    return [
        b.rid,
        b.nickname,
        len({t.speaker_label for t in b.transcripts if t.speaker_label and t.speaker_label != "待确认"}),
        len(b.transcripts),
        round(sum(t.duration_sec or 0 for t in b.transcripts), 1),
        sum(t.char_count for t in b.transcripts),
        len(b.chats),
        b.event_counts.get("member", 0),
        b.event_counts.get("like", 0),
        peak,
        last_pv,
    ]


def summary_csv_bytes(bundles: list[RoomBundle]) -> bytes:
    """生成 summary.csv 的字节流（utf-8-sig，Excel 直接打开不乱码）。"""
    import io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(SUMMARY_HEADER)
    for b in bundles:
        w.writerow(summary_row(b))
    return buf.getvalue().encode("utf-8-sig")


def write_summary(bundles: list[RoomBundle]) -> None:
    (config.EXPORT_DIR / "summary.csv").write_bytes(summary_csv_bytes(bundles))
    wb = build_summary_workbook(bundles)
    try:
        wb.save(config.EXPORT_DIR / "summary.xlsx")
    except PermissionError:
        wb.save(config.EXPORT_DIR / f"summary_{int(time.time())}.xlsx")
    try:
        wb.save(config.EXPORT_DIR / "汇总.xlsx")
    except PermissionError:
        # 用户可能正用 Excel/WPS 打开中文总表；summary.xlsx 已成功写出，不让导出整体失败。
        pass


def export_room_ids(rids: list[str] | None = None) -> list[str]:
    """导出的房间范围：显式参数优先，否则导出数据库历史与当前监听清单的并集。"""
    if rids:
        return sorted({str(r).strip() for r in rids if str(r).strip()})
    return sorted(set(monitored_room_ids()) | set(all_room_ids()))


def export_all(rids: list[str] | None = None, *, cleanup: bool = False) -> list[RoomBundle]:
    """导出指定（或当前监听清单）房间的 md/xlsx/summary，返回各房间数据包。"""
    config.ensure_export_dir()
    ids = export_room_ids(rids)
    if cleanup and ids:
        cleanup_unmonitored_data(ids)
        cleanup_export_files(ids)
    meta = room_display_names()
    speaker_labels = load_speaker_labels()
    bundles = [build_bundle(r, meta.get(r, ""), speaker_labels) for r in ids]
    stems = unique_bundle_stems(bundles)
    for b in bundles:
        stem = stems[b.rid]
        (config.EXPORT_DIR / f"{stem}.md").write_text(render_markdown(b), encoding="utf-8")
        build_workbook(b.rid).save(config.EXPORT_DIR / f"{stem}.xlsx")
    write_summary(bundles)
    return bundles


# ---------- 示例导出（演示数据，全程内存构造，绝不读写任何真实数据库）----------

SAMPLE_RID = "demo"
SAMPLE_NICKNAME = "示例主播（演示数据）"


def _sample_timed_chats(base: int) -> list[tuple[int, str, str]]:
    """演示弹幕：毫秒时间戳，分布在话术时间轴附近。"""
    ms = base * 1000
    raw = [
        (5, "夜风", "主播讲得好细，蹲一个优惠"),
        (12, "小桃", "这个户型采光怎么样？"),
        (20, "阿强", "首付大概多少呀"),
        (33, "Lily", "刚进来，错过了啥"),
        (48, "老王", "能讲讲物业费吗"),
        (61, "momo", "已经在路上了，马上到售楼处"),
        (75, "海带丝", "这价格还能谈吗"),
        (92, "晴天", "关注了，明天来看房"),
    ]
    return [(ms + s * 1000, name, text) for s, name, text in raw]


def sample_bundle() -> RoomBundle:
    """构造一份逼真的演示数据包。时间用当天整点起，导出后时间列可读。"""
    base = int(datetime.now().replace(hour=20, minute=0, second=0, microsecond=0).timestamp())
    speeches = [
        ("各位家人们晚上好，欢迎来到直播间，今天给大家讲解大华锦绣麓城的几款精装户型。", "发言人A", 58.0),
        ("先看这套建面95平的三房，南北通透，主卧带飘窗，采光非常好。", "发言人A", 62.0),
        ("有家人问首付，目前首付方案最低两成，具体以案场为准，可以私信我。", "发言人A", 47.0),
        ("我来补充一下物业这块，物业费每平米两块六，含24小时安保和园林维护。", "发言人B", 55.0),
        ("今晚直播间专属福利：到访就送精美礼品，下定再减三个点，名额有限。", "发言人A", 51.0),
    ]
    transcripts: list[TranscriptRow] = []
    t = base
    for i, (text, spk, dur) in enumerate(speeches, start=1):
        transcripts.append(TranscriptRow(
            room_id=SAMPLE_RID,
            segment_ts=t,
            duration_sec=dur,
            text=text,
            char_count=len(text),
            mp3_name=f"{SAMPLE_RID}/seq{i:05d}.mp3",
            segment_id=i,
            recording_status="ok",
            capture_start=float(t),
            capture_end=float(t) + dur,
            speaker_label=spk,
            speaker_similarity=0.86,
            speaker_change="change_confirmed_start" if spk == "发言人B" else "",
        ))
        t += int(dur) + 3
    timeline = [
        RecordingTimelineRow(
            id=i, room_id=SAMPLE_RID, seq=i, kind="segment", status="ok",
            file_path=tr.mp3_name, capture_start=tr.capture_start, capture_end=tr.capture_end,
            duration_sec=tr.duration_sec, file_size=int((tr.duration_sec or 0) * 16000),
            error=None, transcribed=True,
            transcript_preview=(tr.text[:60] + "…") if len(tr.text) > 60 else tr.text,
        )
        for i, tr in enumerate(transcripts, start=1)
    ]
    timed_chats = _sample_timed_chats(base)
    chats = [(name, content) for _, name, content in timed_chats]
    stats = [(int((base + s) * 1000), online, pv) for s, online, pv in (
        (0, 128, 3400), (30, 156, 3620), (60, 203, 3980),
        (120, 188, 4310), (180, 241, 4675), (240, 262, 5020),
    )]
    event_counts = {"chat": len(chats), "member": 47, "like": 1860, "gift": 9}
    return RoomBundle(SAMPLE_RID, SAMPLE_NICKNAME, transcripts, timeline, chats, stats, event_counts)


def build_sample_workbook():
    """演示工作簿：与真实单房间导出同结构同样式，但数据全在内存，不碰任何库。"""
    from openpyxl import Workbook

    b = sample_bundle()
    timed_chats = _sample_timed_chats(int(b.stats[0][0] // 1000))
    chats_provider = lambda _rid: timed_chats  # noqa: E731  仅样本内联，避免读真库

    wb = Workbook()

    ws = wb.active
    ws.title = "话术"
    ws.append(TRANSCRIPT_HEADER)
    for r in _transcript_rows(b.transcripts):
        ws.append(r)
    _style_sheet(ws, [20, 20, 12, 9, 7, 90, 14, 14, 28], wrap_col=6)

    ws_all = wb.create_sheet("总表")
    ws_all.append(TOTAL_HEADER)
    for row in _total_rows([b], chats_provider):
        ws_all.append(row)
    _style_sheet(
        ws_all,
        [14, 16, 14, 20, 20, 10, 90, 8, 12, 28, 12, 14, 14, 14, 14, 14, 14, 14, 20],
        wrap_col=7,
    )

    ws_speaker = wb.create_sheet("发言人汇总")
    ws_speaker.append(SPEAKER_SUMMARY_HEADER)
    for row in _speaker_summary_rows([b]):
        ws_speaker.append(row)
    _style_sheet(ws_speaker, [14, 16, 14, 12, 16, 12, 20, 20, 14])

    ws_tl = wb.create_sheet("录音时间轴")
    ws_tl.append(TIMELINE_HEADER)
    for r in _timeline_rows(b.timeline):
        ws_tl.append(r)
    _style_sheet(ws_tl, [8, 10, 10, 20, 20, 9, 10, 34, 9, 60, 40], wrap_col=10)

    ws_chat = wb.create_sheet("弹幕")
    ws_chat.append(CHAT_HEADER)
    for r in _chat_rows(timed_chats):
        ws_chat.append(r)
    _style_sheet(ws_chat, [20, 22, 60])

    ws_stat = wb.create_sheet("直播数据")
    ws_stat.append(STAT_HEADER)
    for r in _stat_rows(b.stats):
        ws_stat.append(r)
    _style_sheet(ws_stat, [20, 12, 12])

    return wb


def sample_xlsx_bytes() -> bytes:
    """演示工作簿字节流，供 Web 直接下载/保存。不读写任何真实数据。"""
    import io
    buf = io.BytesIO()
    build_sample_workbook().save(buf)
    return buf.getvalue()


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    config.ensure_export_dir()
    rids = export_room_ids(sys.argv[1:] or None)
    if not rids:
        print("没有可导出的数据（两个库都为空或不存在）。", flush=True)
        return 0
    # 导出默认必须是只读行为。历史清理只能由用户显式触发，不能夹带在导出里。
    cleaned = {"transcripts": 0, "recording_timeline": 0, "events": 0, "room_meta": 0}
    removed_exports = 0
    meta = room_display_names()
    speaker_labels = load_speaker_labels()
    bundles = [build_bundle(r, meta.get(r, ""), speaker_labels) for r in rids]
    stems = unique_bundle_stems(bundles)
    for b in bundles:
        stem = stems[b.rid]
        (config.EXPORT_DIR / f"{stem}.md").write_text(render_markdown(b), encoding="utf-8")
        build_workbook(b.rid).save(config.EXPORT_DIR / f"{stem}.xlsx")
    write_summary(bundles)
    for b in bundles:
        print(f"[OK] {stems[b.rid]}.xlsx/.md  话术{len(b.transcripts)}段/"
              f"弹幕{len(b.chats)}条/在线峰值{max((s[1] for s in b.stats), default=0)}",
              flush=True)
    print(f"清理旧数据: {cleaned}; 清理旧导出文件: {removed_exports}", flush=True)
    print(f"汇总: {config.EXPORT_DIR / 'summary.csv'}  共 {len(bundles)} 房间。", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
