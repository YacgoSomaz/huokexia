from __future__ import annotations

from fastapi.testclient import TestClient


def test_standalone_app_has_a_dedicated_product_boundary() -> None:
    from lead_shrimp.app import PRODUCT_CODE, app, required_feature_for_path

    assert app.title == "获客虾"
    assert PRODUCT_CODE == "lead_shrimp"
    assert required_feature_for_path("/api/comment-leads/run") == "lead_radar"
    assert required_feature_for_path("/api/comment-leads/export") == "export"
    assert required_feature_for_path("/api/license/activate") is None


def test_license_status_is_available_before_card_activation() -> None:
    from lead_shrimp.app import app

    response = TestClient(app).get("/api/license/status")

    assert response.status_code == 200
    payload = response.json()
    assert payload["product_code"] == "lead_shrimp"
    assert "licensed" in payload
    assert payload["licensed"] is False
    assert payload["license_required"] is False


def test_health_endpoint_is_independent_from_license_status() -> None:
    from lead_shrimp.app import app

    response = TestClient(app).get("/api/health")

    assert response.status_code == 200
    assert response.json() == {"ok": True, "service": "lead_shrimp"}


def test_frontend_does_not_expose_replay_navigation() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert "AI 获客系统" not in page
    assert "评论线索工作台" in page
    assert "/api/license/activate" in page


def test_launcher_removes_its_browser_only_flag_before_starting_api() -> None:
    from lead_shrimp.launcher import app_argv

    assert app_argv(["LeadShrimpLauncher.exe", "--port", "8922", "--no-browser"]) == [
        "LeadShrimpLauncher.exe",
        "--port",
        "8922",
    ]


def test_launcher_rewrites_the_requested_port_for_the_selected_free_port() -> None:
    from lead_shrimp.launcher import with_port

    assert with_port(["LeadShrimpLauncher.exe", "--port", "8922"], 8931) == [
        "LeadShrimpLauncher.exe",
        "--port",
        "8931",
    ]
    assert with_port(["LeadShrimpLauncher.exe"], 8931) == [
        "LeadShrimpLauncher.exe",
        "--port",
        "8931",
    ]


def test_launcher_falls_back_when_the_preferred_port_is_busy(monkeypatch) -> None:
    from lead_shrimp import launcher

    attempts: list[int] = []

    def fake_bind(port: int) -> bool:
        attempts.append(port)
        return port != 8922

    monkeypatch.setattr(launcher, "_port_is_free", fake_bind)

    assert launcher.choose_port(8922, attempts=3) == 8923
    assert attempts == [8922, 8923]


def test_launcher_waits_for_the_health_endpoint_before_opening_browser(monkeypatch) -> None:
    from lead_shrimp import launcher

    opened: list[str] = []
    monkeypatch.setattr(launcher.webbrowser, "open", lambda url, new=0: opened.append(url) or True)
    responses = iter([False, True])
    monkeypatch.setattr(launcher, "service_ready", lambda _port: next(responses))

    assert launcher.wait_for_service_and_open(8931, timeout_sec=1, poll_sec=0) is True
    assert opened == ["http://127.0.0.1:8931/"]


def test_launcher_health_probe_uses_the_dedicated_health_endpoint() -> None:
    from lead_shrimp import launcher

    source = __import__("inspect").getsource(launcher.service_ready)

    assert "/api/health" in source


def test_launcher_records_startup_exceptions_for_pythonw_users() -> None:
    from lead_shrimp import launcher

    source = __import__("inspect").getsource(launcher.main)

    assert "traceback" in source
    assert "_startup_error_file" in source


def test_public_installer_and_startup_page_use_simplified_chinese() -> None:
    from pathlib import Path
    from lead_shrimp import launcher

    iss = Path(__file__).parents[1] / "build" / "lead_shrimp_public.iss"
    installer_source = iss.read_text(encoding="utf-8")
    launcher_source = __import__("inspect").getsource(launcher._startup_error_file)

    assert 'Name: "chinesesimplified"' in installer_source
    assert "ChineseSimplified.isl" in installer_source
    assert "服务启动失败，请查看日志" in launcher_source
    assert "traceback" not in launcher_source.lower()


def test_public_build_bundles_playwright_chromium_fallback() -> None:
    from pathlib import Path

    build_source = (Path(__file__).parents[1] / "build" / "build_public_release.ps1").read_text(encoding="utf-8")
    launcher_source = __import__("inspect").getsource(__import__("lead_shrimp.launcher", fromlist=["main"]).main)

    assert "PLAYWRIGHT_BROWSERS_PATH" in build_source
    assert "playwright install chromium" in build_source
    assert "PLAYWRIGHT_BROWSERS_PATH" in launcher_source


def test_browser_launch_error_keeps_attempt_details() -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads._launch_comment_context)

    assert "attempt_errors" in source
    assert "msedge" in source
    assert "chromium" in source


def test_public_installer_handles_running_app_and_locked_files() -> None:
    from pathlib import Path

    iss = Path(__file__).parents[1] / "build" / "lead_shrimp_public.iss"
    installer_source = iss.read_text(encoding="utf-8")

    assert "CloseApplications=yes" in installer_source
    assert "RestartApplications=no" in installer_source
    assert "CloseApplicationsFilter=" in installer_source
    assert "python\\pythonw.exe" in installer_source
    assert "node.exe" in installer_source


def test_monitor_api_supports_delete_without_deleting_historical_leads() -> None:
    from pipeline import comment_leads
    from lead_shrimp import app as app_module

    store = {
        "version": 1,
        "monitors": [{"id": "monitor-a", "title": "账号 A"}, {"id": "monitor-b"}],
        "leads": [{"lead_id": "lead-a", "monitor_id": "monitor-a"}],
        "jobs": [{"id": "job-a", "monitor_id": "monitor-a"}],
    }
    saved: list[dict] = []
    monkeypatch = __import__("pytest").MonkeyPatch()
    try:
        monkeypatch.setattr(comment_leads, "load_store", lambda: store)
        monkeypatch.setattr(comment_leads, "save_store", lambda value: saved.append(value))
        response = app_module.api_comment_leads_delete_monitor("monitor-a")
    finally:
        monkeypatch.undo()
    assert response.body == b'{"ok":true,"deleted":true,"monitor_id":"monitor-a"}'
    assert [row["id"] for row in store["monitors"]] == ["monitor-b"]
    assert store["leads"] == [{"lead_id": "lead-a", "monitor_id": "monitor-a"}]
    assert store["jobs"] == []
    assert saved


def test_login_endpoint_short_circuits_when_saved_login_is_valid(monkeypatch) -> None:
    from pipeline import comment_leads
    from lead_shrimp import app as app_module

    monkeypatch.setattr(comment_leads, "login_status", lambda: {"logged_in": True, "browser": "msedge"})
    monkeypatch.setattr(comment_leads, "open_login_browser", lambda **_: (_ for _ in ()).throw(AssertionError("must not open")))
    response = app_module.api_comment_leads_login({})
    assert response.status_code == 200
    assert __import__("json").loads(response.body)["already_logged_in"] is True


def test_login_and_capture_window_arguments_restore_login_window(monkeypatch) -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads._launch_comment_context)
    assert "--window-position=80,80" in source
    assert "--window-size=1365,768" in source
    assert "--window-position=-32000,-32000" in source
    assert "background" in source


def test_login_browser_triggers_the_visible_scan_login_panel() -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads._trigger_login_panel)
    assert "扫码登录" in source
    assert "登录" in source
    assert "placeholder" in source
    assert "keyboard.press(\"Enter\")" in source


def test_persistent_profile_login_cookies_are_not_overwritten_by_shared_cache() -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads._seed_shared_jar_if_needed)
    assert "_context_cookie_jar(context)" in source
    assert "short_video._has_douyin_login_cookie" in source
    assert "context.add_cookies" in source


def test_frontend_does_not_reopen_login_window_when_already_logged_in() -> None:
    from pathlib import Path

    page = (Path(__file__).parents[1] / "frontend.html").read_text(encoding="utf-8")
    assert "抖音已登录，无需重复打开登录窗口" in page
    assert "data-monitor-delete" in page
    assert "/api/comment-leads/monitors/" in page


def test_frontend_uses_competitor_accounts_label_without_advanced_entry() -> None:
    from pathlib import Path

    page = (Path(__file__).parents[1] / "frontend.html").read_text(encoding="utf-8")
    assert "对标账号" in page
    assert "高级入口" not in page
    assert "管理监控账号" not in page


def test_frontend_places_lead_time_before_commenter_name() -> None:
    from pathlib import Path

    page = (Path(__file__).parents[1] / "frontend.html").read_text(encoding="utf-8")
    assert '<span>时间</span><span>评论人</span>' in page
    assert page.index("formatTime(x.create_time)") < page.index('class="person"')


def test_frontend_improves_date_picker_and_primary_return_action() -> None:
    from pathlib import Path

    page = (Path(__file__).parents[1] / "frontend.html").read_text(encoding="utf-8")
    assert "showPicker" in page
    assert 'id="dragSelectLeads"' not in page
    assert 'class="btn return-workbench"' in page


def test_frontend_removes_dashboard_metrics_and_keeps_status_in_workflow() -> None:
    from pathlib import Path

    page = (Path(__file__).parents[1] / "frontend.html").read_text(encoding="utf-8")
    assert "今天的线索看板" not in page
    assert 'id="loginMetric"' not in page
    assert 'id="competitorCount"' in page
    assert 'id="sideLeadCount"' in page
    assert "请检查抖音登录状态" in page


def test_competitor_cards_are_compact_and_use_collect_works_action() -> None:
    from pathlib import Path

    page = (Path(__file__).parents[1] / "frontend.html").read_text(encoding="utf-8")
    assert "monitor-link" not in page
    assert "选择作品</button>" not in page
    assert "采集作品</button>" in page
    assert ".monitor-actions{display:flex" in page


def test_app_disables_uvicorn_console_log_config_for_pythonw() -> None:
    from lead_shrimp import app

    source = __import__("inspect").getsource(app.main)

    assert "log_config=None" in source


def test_collection_failure_is_returned_as_json_for_the_frontend(monkeypatch) -> None:
    from lead_shrimp import app as app_module

    monkeypatch.setattr(app_module.comment_leads, "run_monitor", lambda _monitor_id: (_ for _ in ()).throw(RuntimeError("浏览器连接已断开")))
    response = TestClient(app_module.app).post("/api/comment-leads/run", json={"monitor_id": "monitor-test"})

    assert response.status_code == 503
    assert response.headers["content-type"].startswith("application/json")
    assert response.json()["detail"] == "评论采集失败：浏览器连接已断开"


def test_frontend_explains_partial_collection_and_non_json_errors() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert "服务端返回了无法识别的数据" in page
    assert "部分完成" in page


def test_frontend_only_auto_opens_activation_when_license_is_required() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert "const required=Boolean(data.license_required ?? data.enforced);" in page
    assert "if(!ok&&required)$('licenseModal').classList.add('show');" in page
    assert "本地模式可直接使用" in page


def test_collection_diagnosis_turns_missing_login_into_a_next_step(monkeypatch) -> None:
    from lead_shrimp import app as app_module

    monkeypatch.setattr(app_module.comment_leads, "list_monitors", lambda: [{"id": "monitor-a", "last_count": 20, "last_error": "未登录抖音，请先点击授权登录", "last_run_at": 123, "discovered_video_count": 5}])
    monkeypatch.setattr(app_module.comment_leads, "login_status", lambda: {"logged_in": False, "has_profile": True, "cookie_count": 0})

    response = TestClient(app_module.app).get("/api/comment-leads/diagnosis?monitor_id=monitor-a")

    assert response.status_code == 200
    assert response.json()["next_action"] == "login"
    assert response.json()["title"] == "需要完成抖音登录"
    assert response.json()["captured"] == 20


def test_comment_login_state_survives_page_refresh(monkeypatch, tmp_path) -> None:
    from lead_shrimp import app as app_module

    profile = tmp_path / "comment_profile"
    profile.mkdir()
    (profile / "Default").mkdir()
    state_path = tmp_path / "comment_login_state.json"
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_PROFILE_DIR", profile)
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_LOGIN_STATE_JSON", state_path)
    monkeypatch.setattr(app_module.comment_leads.browser_cookies, "shared_status", lambda: {"has_login": False, "cookie_count": 0, "browser": "msedge"})

    app_module.comment_leads._save_login_state(True, "msedge")
    first = TestClient(app_module.app).get("/api/comment-leads/status").json()
    second = TestClient(app_module.app).get("/api/comment-leads/status").json()

    assert first["logged_in"] is True
    assert second["logged_in"] is True
    assert first["expires_at"] > first.get("cookie_count", 0)


def test_login_waits_until_authenticated_without_a_short_deadline() -> None:
    from lead_shrimp import app as app_module
    from pipeline import comment_leads

    app_source = __import__("inspect").getsource(app_module.api_comment_leads_login)
    login_source = __import__("inspect").getsource(comment_leads.open_login_browser)
    blocking_source = __import__("inspect").getsource(app_module._run_blocking)

    assert "default=0" in app_source
    assert "_process_timeout_sec=None" in app_source
    assert "timeout_sec = None if wait_ms <= 0 else" in login_source
    assert "proc.join(_process_timeout_sec)" in blocking_source


def test_expired_comment_login_state_requires_relogin(monkeypatch, tmp_path) -> None:
    from lead_shrimp import app as app_module

    profile = tmp_path / "comment_profile"
    profile.mkdir()
    (profile / "Default").mkdir()
    state_path = tmp_path / "comment_login_state.json"
    state_path.write_text('{"authenticated": true, "browser": "msedge", "updated_at": 1, "expires_at": 2}', encoding="utf-8")
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_PROFILE_DIR", profile)
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_LOGIN_STATE_JSON", state_path)
    monkeypatch.setattr(app_module.comment_leads.browser_cookies, "shared_status", lambda: {"has_login": False, "cookie_count": 0, "browser": "msedge"})

    response = TestClient(app_module.app).get("/api/comment-leads/status")

    assert response.status_code == 200
    assert response.json()["logged_in"] is False


def test_frontend_has_a_fast_diagnosis_area() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert "快速排错" in page
    assert "/api/comment-leads/diagnosis" in page


def test_frontend_uses_a_compact_workbench_and_visible_controls() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert ".app{max-width:none" in page
    assert ".tabs{width:184px" in page
    assert ".tab-btn{display:flex;align-items:center" in page
    assert ".tab-btn.active{background:#fff" in page


def test_completed_setup_status_bar_collapses_to_reduce_top_blank_space() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert ".setup.is-complete{display:none}" in page
    assert "$('setupPanel').classList.toggle('is-complete'" in page


def test_frontend_groups_lead_filters_into_a_compact_workbench_toolbar() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert '<div class="toolbar lead-toolbar">' in page
    assert 'class="date-range"' in page
    assert 'class="selection-tools"' in page
    assert ".lead-toolbar{" in page
    assert ".date-range{" in page
    assert ".selection-tools{" in page


def test_frontend_exposes_task_oriented_navigation_and_selection_feedback() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'class="tab-index"' in page
    assert 'id="selectionSummary"' in page
    assert ".tabs-caption{" in page
    assert ".lead-selection-summary{" in page


def test_frontend_starts_with_setup_and_orders_tabs_by_real_workflow() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    add = page.index('data-tab="addTab"')
    works = page.index('data-tab="worksTab"')
    monitors = page.index('data-tab="monitorsTab"')
    assert add < works < monitors
    assert 'data-tab="leadsTab"' not in page
    assert "function chooseInitialTab()" in page
    assert "chooseInitialTab" in page
    assert "function updateWorkflowNav()" in page


def test_frontend_keeps_comments_out_of_the_primary_flow_and_opens_them_on_demand() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="closeLeads"' in page
    assert 'class="tab-panel leads-overlay' in page
    assert 'function openLeadsPanel()' in page
    assert 'data-step="login"' in page
    assert 'data-tab="leadsTab"' not in page.split('<nav class="tabs"', 1)[1].split('</nav>', 1)[0]


def test_frontend_exposes_a_sidebar_lead_button_and_cached_monitor_management() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="openLeadsSidebar"' in page
    assert 'id="sideLeadCount"' in page
    assert 'leads-shortcut' in page
    assert "querySelectorAll('[data-tab]')" in page
    assert 'class="monitor-meta"' in page
    assert 'works_cached_at' in page
    assert 'cached_videos.length' in page


def test_work_picker_defaults_to_recent_non_pinned_videos() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="workAgeFilter"' in page
    assert 'id="includePinnedWorks"' in page
    assert "const DEFAULT_WORK_WINDOW_DAYS=30;" in page
    assert "v.selected===true?'checked':''" in page


def test_profile_render_waits_for_api_metadata_before_dom_fallback() -> None:
    from pipeline.short_video import _render_profile_events_locked

    source = __import__("inspect").getsource(_render_profile_events_locked)

    assert "tick >= 14" in source


def test_profile_fallback_enriches_video_ids_from_detail_endpoint() -> None:
    from pipeline import short_video

    source = __import__("inspect").getsource(short_video._enrich_profile_videos)

    assert "aweme/detail" in source
    assert "aweme_id" in source


def test_incomplete_profile_cache_does_not_short_circuit_metadata_refresh() -> None:
    from pipeline import short_video

    source = __import__("inspect").getsource(short_video.resolve_profile)

    assert "_profile_video_has_metadata" in source


def test_work_picker_shows_pinned_works_and_marks_them_without_selecting() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert '<option value="0" selected>全部作品</option>' in page
    assert "$('includePinnedWorks').checked=true;" in page
    assert "${v.pinned?'<span class=\"chip warn\">置顶</span>':''}" in page


def test_profile_metadata_retries_captured_post_endpoint() -> None:
    from pipeline.short_video import _render_profile_events_locked

    source = __import__("inspect").getsource(_render_profile_events_locked)

    assert "metadata_retry" in source
    assert "dom_fallback_videos" in source


def test_profile_post_endpoint_retries_are_bounded() -> None:
    from pipeline.short_video import _render_profile_events_locked

    source = __import__("inspect").getsource(_render_profile_events_locked)

    assert "post_api_attempts" in source
    assert "post_api_attempts[api_url]" in source


def test_profile_api_requests_have_a_per_run_budget() -> None:
    from pipeline.short_video import _render_profile_events_locked

    source = __import__("inspect").getsource(_render_profile_events_locked)

    assert "profile_api_request_count" in source
    assert "profile_api_request_count >= 4" in source
    assert "post_api_urls[-2:]" in source


def test_author_comments_are_removed_before_lead_ingestion() -> None:
    from pipeline.comment_leads import filter_author_comments

    rows = [
        {"comment_id": "author", "commenter_sec_uid": "sec-author", "content": "作者评论"},
        {"comment_id": "visitor", "commenter_sec_uid": "sec-visitor", "content": "用户评论"},
    ]

    filtered = filter_author_comments(rows, author_sec_uid="sec-author")

    assert [row["comment_id"] for row in filtered] == ["visitor"]


def test_profile_work_cache_is_used_before_browser_resolution() -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads.resolve_profile_works)

    assert "cached_videos" in source
    assert "force" in source
    assert "已使用本地作品缓存" in source


def test_frontend_has_lead_time_filters_and_cached_work_refresh() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="leadAgeFilter"' in page
    assert 'id="leadSort"' in page
    assert 'data-refresh="' in page
    assert "payload.force=true" in page


def test_selected_lead_export_only_accepts_existing_ids(tmp_path, monkeypatch) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from lead_shrimp import app as app_module

    export_dir = tmp_path / "exports"
    store_path = tmp_path / "leads.json"
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_JSON", store_path)
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_EXPORT_DIR", export_dir)
    store_path.write_text(
        '{"version": 1, "monitors": [], "leads": ['
        '{"lead_id":"lead-a","comment_id":"comment-a","content":"高价值"},'
        '{"lead_id":"lead-b","comment_id":"comment-b","content":"普通"}], "jobs": []}',
        encoding="utf-8",
    )

    response = TestClient(app_module.app).post(
        "/api/comment-leads/export", json={"lead_ids": ["lead-a"]}
    )

    assert response.status_code == 200
    values = list(load_workbook(BytesIO(response.content)).active.values)
    assert any("高价值" in row for row in values)
    assert not any("普通" in row for row in values)


def test_selected_lead_export_is_a_compact_chinese_xlsx(tmp_path, monkeypatch) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    from lead_shrimp import app as app_module

    export_dir = tmp_path / "exports"
    store_path = tmp_path / "leads.json"
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_JSON", store_path)
    monkeypatch.setattr(app_module.comment_leads.config, "COMMENT_LEADS_EXPORT_DIR", export_dir)
    store_path.write_text(
        '{"version": 1, "monitors": [], "leads": ['
        '{"lead_id":"lead-a","create_time":1735498402,"commenter_nickname":"张三",'
        '"content":"想了解价格和位置","comment_ip_location":"云南",'
        '"commenter_profile_url":"https://www.douyin.com/user/sec-a","status":"待联系"}'
        '], "jobs": []}',
        encoding="utf-8",
    )

    response = TestClient(app_module.app).post(
        "/api/comment-leads/export", json={"lead_ids": ["lead-a"]}
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].endswith(".xlsx")
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == ["评论时间", "评论人", "评论内容", "评论人IP", "评论人主页", "联系状态"]
    assert [cell.value for cell in sheet[2]][1:] == ["张三", "想了解价格和位置", "云南", "https://www.douyin.com/user/sec-a", "待联系"]
    assert sheet.column_dimensions["A"].width >= 18
    assert sheet.column_dimensions["C"].width >= 40
    assert sheet.row_dimensions[2].height >= 30
    assert sheet.freeze_panes == "A2"


def test_frontend_supports_selective_lead_export() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="selectVisibleLeads"' in page
    assert 'data-lead-id=' in page
    assert "导出选中" in page
    assert "/api/comment-leads/export" in page


def test_comment_collection_default_is_not_limited_to_100() -> None:
    from lead_shrimp.app import frontend_path
    from pipeline import comment_leads

    page = frontend_path().read_text(encoding="utf-8")
    source = __import__("inspect").getsource(comment_leads.capture_video_comments)

    assert '<option value="500" selected>500 条</option>' in page
    assert "max_comments: int = 500" in source
    assert "max_capture_seconds = min(300" in source


def test_work_selection_has_a_comment_count_control() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="workCommentLimit"' in page
    assert 'option value="500" selected' in page
    assert "$('workCommentLimit').value" in page


def test_cached_works_restore_after_refresh_and_warn_after_one_day() -> None:
    from lead_shrimp.app import frontend_path
    from pipeline import comment_leads

    page = frontend_path().read_text(encoding="utf-8")
    source = __import__("inspect").getsource(comment_leads._profile_monitor_metadata)

    assert "works_cached_at" in source
    assert "function loadCachedWorks" in page
    assert "超过 1 天未刷新作品" in page
    assert "cached_videos" in page


def test_profile_work_refresh_marks_new_videos_and_comment_increase() -> None:
    from pipeline import comment_leads

    current = [
        {"id": "new", "comment_count": 4},
        {"id": "old", "comment_count": 18},
    ]
    previous = [{"id": "old", "comment_count": 12}]

    marked = comment_leads.annotate_profile_video_changes(current, previous)

    assert marked[0]["is_new"] is True
    assert marked[0]["comment_increase"] == 4
    assert marked[1]["is_new"] is False
    assert marked[1]["comment_increase"] == 6


def test_work_picker_supports_competitor_account_filter_and_change_badges() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="workCompetitorFilter"' in page
    assert "按对标账号查看作品" in page
    assert "work-new-badge" in page
    assert "新视频" in page
    assert "work-comment-badge" in page
    assert "comment_increase" in page


def test_work_picker_uses_comment_collection_action_with_progress_feedback() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="collectSelectedWorks"' in page
    assert '>采集评论<' in page
    assert 'id="commentProgress"' in page
    assert 'id="commentProgressBar"' in page
    assert "setCommentProgress(true" in page
    assert "setCommentProgress(false" in page


def test_work_picker_shows_zero_comment_works_but_blocks_collection() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert "comment_count===0" in page
    assert "评论为 0" in page
    assert "请刷新作品信息后再试" in page
    assert 'id="refreshWorksInfo"' in page


def test_profile_work_count_defaults_to_five_and_caps_at_twenty() -> None:
    from lead_shrimp.app import frontend_path
    from lead_shrimp import app as app_module
    from pipeline import comment_leads

    page = frontend_path().read_text(encoding="utf-8")

    assert '<option value="5">5 条</option>' in page
    assert '<option value="20">20 条</option>' in page
    assert 'maximum=20' in __import__("inspect").getsource(app_module.api_comment_leads_profile_videos)
    assert 'min(int(max_videos or 5), 20)' in __import__("inspect").getsource(comment_leads.resolve_profile_works)


def test_ingested_comment_keeps_a_minimal_video_context_snapshot(monkeypatch) -> None:
    from pipeline import comment_leads

    store = {"version": 1, "monitors": [], "leads": [], "jobs": []}
    monkeypatch.setattr(comment_leads, "load_store", lambda: store)
    monkeypatch.setattr(comment_leads, "save_store", lambda _value: None)

    comment_leads.ingest_rows(
        [{"comment_id": "comment-1", "aweme_id": "video-1", "content": "附近有公园吗"}],
        monitor_id="monitor-1",
        video_context={"id": "video-1", "title": "昆明精装现房", "publish_time": 123, "like_count": 8, "comment_count": 3, "pinned": False},
    )

    assert store["leads"][0]["video_context"] == {
        "id": "video-1", "title": "昆明精装现房", "publish_time": 123,
        "like_count": 8, "comment_count": 3, "pinned": False,
    }


def test_same_commenter_leads_are_grouped_with_all_concerns() -> None:
    from pipeline import comment_leads

    groups = comment_leads.group_leads_by_commenter([
        {"lead_id": "a", "commenter_sec_uid": "u1", "commenter_nickname": "甲", "content": "附近有公园吗", "create_time": 100},
        {"lead_id": "b", "commenter_sec_uid": "u1", "commenter_nickname": "甲", "content": "孩子怎么上学", "create_time": 200},
        {"lead_id": "c", "commenter_sec_uid": "u2", "commenter_nickname": "乙", "content": "多少钱一平", "create_time": 150},
    ])

    assert [group["commenter_key"] for group in groups] == ["u1", "u2"]
    assert groups[0]["comment_count"] == 2
    assert [item["content"] for item in groups[0]["comments"]] == ["孩子怎么上学", "附近有公园吗"]


def test_internal_test_api_is_isolated_from_production_lead_routes(monkeypatch) -> None:
    from lead_shrimp import app as app_module

    monkeypatch.setattr(app_module.comment_leads, "list_internal_test_overview", lambda: {"mode": "internal_test", "tasks": [], "groups": []})
    response = TestClient(app_module.app).get("/api/comment-leads/internal-test/overview")

    assert response.status_code == 200
    assert response.json()["mode"] == "internal_test"


def test_frontend_has_an_internal_test_tab_without_replacing_production_flow() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'data-tab="internalTestTab"' in page
    assert 'id="internalTestTab"' in page
    assert '内部测试中' in page
    assert 'internal-test/overview' in page
    assert 'data-ai-agent="overview"' in page
    assert 'data-ai-agent="scanner"' in page
    assert 'data-ai-agent="screening"' in page
    assert 'data-ai-agent="opportunity"' in page
    assert 'function renderInternalAgent' in page


def test_frontend_renders_ai_agent_choices_as_clear_button_options() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert ".ai-agent{width:100%;text-align:left;border:1px solid" in page
    assert ".ai-agent:hover{" in page
    assert ".ai-agent.active{background:#eaf2ff" in page


def test_internal_daily_scan_only_collects_changed_videos(monkeypatch) -> None:
    from pipeline import comment_leads

    store = {"version": 1, "monitors": [{"id": "m1", "target_type": "profile", "target_url": "https://example.test/user/a", "max_videos": 5, "max_comments": 100}], "leads": [], "jobs": [], "internal_test_tasks": []}
    monkeypatch.setattr(comment_leads, "load_store", lambda: store)
    monkeypatch.setattr(comment_leads, "save_store", lambda _value: None)
    monkeypatch.setattr(comment_leads, "resolve_profile_works", lambda *_args, **_kwargs: {"videos": [{"id": "v-new", "url": "https://example.test/video/v-new", "is_new": True, "comment_count": 2}, {"id": "v-old", "url": "https://example.test/video/v-old", "comment_increase": 0, "comment_count": 4}]})
    captured: list[list[str]] = []
    monkeypatch.setattr(comment_leads, "run_selected_videos", lambda _mid, videos, **_kwargs: captured.append([v["id"] for v in videos]) or {"captured": 2, "inserted": 2, "error": ""})

    result = comment_leads.run_internal_daily_scan()

    assert result["monitors"] == 1
    assert captured == [["v-new"]]
    assert store["internal_test_tasks"][0]["kind"] == "daily_incremental_scan"


def test_ai_pipeline_validates_results_before_writing_to_leads(monkeypatch) -> None:
    from pipeline import lead_ai

    leads = [{"lead_id": "l1", "content": "附近有公园吗", "create_time": 1, "comment_ip_location": "云南", "video_context": {"title": "昆明现房"}}]
    replies = iter([
        '{"schema_version":"1.0","results":[{"lead_id":"l1","keep":true,"intent_level":"medium","intent_tags":["生活配套"],"summary":"关注公园配套。","reason":"评论询问附近公园。","confidence":0.8,"needs_human_review":false}]}',
        '{"schema_version":"1.0","results":[{"lead_id":"l1","priority":"P2","follow_up_channel":"抖音私信","recommended_action":"72小时内跟进","opening_message":"您更关注公园还是通勤？","rationale":"存在生活配套关注。","risks":[],"confidence":0.75,"needs_human_review":false}]}'
    ])
    monkeypatch.setattr(lead_ai, "_complete_json", lambda *_args, **_kwargs: next(replies))

    result = lead_ai.analyze_leads(leads, business_context={"offer": "昆明现房"})

    assert result["processed"] == 1
    assert leads[0]["ai"]["screening"]["intent_tags"] == ["生活配套"]
    assert leads[0]["ai"]["opportunity"]["priority"] == "P2"


def test_ai_pipeline_rejects_unknown_or_invalid_results(monkeypatch) -> None:
    from pipeline import lead_ai

    monkeypatch.setattr(lead_ai, "_complete_json", lambda *_args, **_kwargs: '{"schema_version":"1.0","results":[{"lead_id":"unknown","keep":true}]}')

    result = lead_ai.analyze_leads([{"lead_id": "l1", "content": "你好"}], business_context={})

    assert result["processed"] == 0
    assert result["errors"]


def test_internal_schedule_runs_once_per_day_after_configured_time(monkeypatch) -> None:
    from pipeline import comment_leads
    store = {"version": 1, "monitors": [], "leads": [], "jobs": [], "internal_test_tasks": [], "internal_test_schedule": {"enabled": True, "time": "08:00", "last_run_date": ""}}
    monkeypatch.setattr(comment_leads, "load_store", lambda: store)
    monkeypatch.setattr(comment_leads, "save_store", lambda _value: None)
    monkeypatch.setattr(comment_leads, "run_internal_daily_scan", lambda: {"ok": True})

    assert comment_leads.run_due_internal_daily_scan("2026-07-17 08:01") is True
    assert comment_leads.run_due_internal_daily_scan("2026-07-17 08:02") is False


def test_comment_capture_is_headless_until_verification_is_detected() -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads.capture_video_comments)

    assert "headed: bool = True" in source
    assert "background: bool = True" in source
    assert "_page_needs_verification" in source
    assert "allow_interactive_fallback" in source


def test_background_capture_does_not_change_login_window_launch() -> None:
    from pipeline import comment_leads

    source = __import__("inspect").getsource(comment_leads._launch_comment_context)

    assert "background" in source
    assert "--start-minimized" in source
    assert "--window-position=-32000,-32000" in source


def test_lead_list_supports_drag_selection_and_date_range_selection() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="leadStartDate"' in page
    assert 'id="leadEndDate"' in page
    assert "selectLeadsByDate" in page
    assert "leadRangeSlider" in page


def test_drag_selection_uses_one_vertical_range_slider_before_checkbox() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert ".lead-range-rail" in page
    assert 'id="leadRangeSlider"' in page
    assert 'id="leadRangeThumb"' in page
    assert 'id="leadRangeRail"' in page
    assert "selectLeadsToIndex" in page
    assert "onpointermove" in page
    assert "style.top" in page
    assert "rows.forEach((row,i)=>setLeadSelection(row.lead_id,i<=end&&!state.rangeExcludedIds.has(String(row.lead_id||''))))" in page
    assert ".lead-range-rail{position:absolute;left:10px;top:42px;bottom:0;width:36px" in page
    assert ".lead-range-slider{position:relative;width:36px" in page
    assert "dragSelectMode:true" in page
    assert "applyRangeToRows" in page
    assert "rowIndexAtY" in page
    assert "getBoundingClientRect" in page
    assert "event.clientY" in page
    assert "setPointerCapture" in page
    assert "requestAnimationFrame" in page
    assert "scrollHeight" in page
    assert "style.height" in page
    assert "rangeExcludedIds" in page
    assert "rangeExcludedIds.add" in page
    assert "!state.rangeExcludedIds.has" in page
    assert "dragSelectMode:true" in page


def test_monitor_list_is_a_compact_avatar_grid() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert ".monitor-list{display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr))" in page
    assert "class=\"monitor-identity\"" in page
    assert "class=\"monitor-avatar\"" in page
    assert 'const url=escapeHtml(m.target_url||m.raw_url||'');' not in page
    assert '作品 ${escapeHtml(m.max_videos||5)} 条' not in page


def test_work_resolution_has_visible_progress_before_request_finishes() -> None:
    from lead_shrimp.app import frontend_path

    page = frontend_path().read_text(encoding="utf-8")

    assert 'id="workLoading"' in page
    assert 'id="workLoadingText"' in page
    assert "function setWorkLoading" in page
    assert "switchTab('worksTab');" in page
    assert "$('resolveWorksBtn').disabled=active;" in page
